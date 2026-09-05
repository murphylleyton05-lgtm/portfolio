"""
Genera Pipeline_ETL_Dataset.xlsx con los problemas de calidad intencionales
del checkpoint (duplicados, nulos y tipos por depurar). Las celdas problematicas
se marcan en AMARILLO, tal como las entrega el curso.

Conteos crudos -> resultado esperado tras limpiar:
  clientes : 12 filas (11 unicos + 1 duplicado, +2 nulos)  -> 11
  productos: 13 filas (12 unicos + 1 duplicado, +2 nulos)  -> 12
  ventas   : 50 filas (limpia)                             -> 50
  categorias: 4 filas (limpia)                             -> 4
"""
import random
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

random.seed(2024)

YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER = Font(bold=True)


def write_sheet(wb, name, headers, rows, highlight):
    """highlight: set de tuplas (fila_datos_0based, col_0based) a pintar de amarillo."""
    ws = wb.create_sheet(title=name)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = HEADER
    for r_idx, row in enumerate(rows):
        ws.append(row)
        for c_idx in range(len(headers)):
            if (r_idx, c_idx) in highlight:
                ws.cell(row=r_idx + 2, column=c_idx + 1).fill = YELLOW
    for c_idx, h in enumerate(headers):
        maxlen = max([len(str(h))] + [len(str(row[c_idx])) for row in rows if row[c_idx] is not None])
        ws.column_dimensions[chr(65 + c_idx)].width = min(maxlen + 3, 40)
    return ws


wb = Workbook()
wb.remove(wb.active)

# ---------------------------------------------------------------- categorias
cat_headers = ["id_categoria", "nombre_categoria"]
cat_rows = [
    [1, "Notebooks"],
    [2, "Perifericos"],
    [3, "Componentes"],
    [4, "Smartphones"],
]
write_sheet(wb, "categorias", cat_headers, cat_rows, set())

# ---------------------------------------------------------------- clientes
cli_headers = ["id_cliente", "nombre_cliente", "email", "ciudad", "pais",
               "fecha_registro", "canal"]
cli_rows = [
    [1,  "Martin Gomez",     "mgomez@mail.com",     "Buenos Aires", "Argentina", date(2023, 1, 15), "Online"],
    [2,  "Lucia Fernandez",  "lfernandez@mail.com", "Cordoba",      "Argentina", date(2023, 2, 3),  "Tienda"],
    [3,  "Carlos Rojas",     None,                  "Santiago",     "Chile",     date(2023, 2, 20), "Online"],   # email nulo
    [4,  "Ana Torres",       "atorres@mail.com",    None,           "Mexico",    date(2023, 3, 11), "Online"],   # ciudad nula
    [5,  "Diego Morales",    "dmorales@mail.com",   "Lima",         "Peru",      date(2023, 4, 5),  "Tienda"],
    [6,  "Sofia Castro",     "scastro@mail.com",    "Bogota",       "Colombia",  date(2023, 5, 19), "Online"],
    [7,  "Javier Nunez",     "jnunez@mail.com",     "Montevideo",   "Uruguay",   date(2023, 6, 22), "Tienda"],
    [8,  "Valentina Ruiz",   "vruiz@mail.com",      "Guadalajara",  "Mexico",    date(2023, 7, 30), "Online"],
    [9,  "Mateo Silva",      "msilva@mail.com",     "Rosario",      "Argentina", date(2023, 8, 14), "Online"],
    [10, "Camila Vargas",    "cvargas@mail.com",    "Quito",        "Ecuador",   date(2023, 9, 25), "Tienda"],
    [11, "Nicolas Herrera",  "nherrera@mail.com",   "Asuncion",     "Paraguay",  date(2023, 10, 8), "Online"],
    [3,  "Carlos Rojas",     None,                  "Santiago",     "Chile",     date(2023, 2, 20), "Online"],   # DUPLICADO de id 3
]
cli_highlight = {
    (2, 2),   # email nulo (fila id=3)
    (3, 3),   # ciudad nula (fila id=4)
    (11, 0), (11, 1), (11, 2), (11, 3), (11, 4), (11, 5), (11, 6),  # fila duplicada
}
write_sheet(wb, "clientes", cli_headers, cli_rows, cli_highlight)

# ---------------------------------------------------------------- productos
prod_headers = ["id_producto", "nombre_producto", "categoria", "precio",
                "costo", "stock", "activo"]
prod_rows = [
    [101, "Notebook Lenovo IdeaPad", "Notebooks",   850.00, 620.00, 15, 1],
    [102, "Notebook HP Pavilion",    "Notebooks",   920.50, 680.00, 10, 1],
    [103, "Mouse Logitech M170",     "Perifericos", 12.99,  6.50,  120, 1],
    [104, "Teclado Redragon K552",   "Perifericos", 45.00,  25.00,  60, 1],
    [105, "Monitor Samsung 24",      "Perifericos", 189.99, 140.00, 25, 1],
    [106, "SSD Kingston 480GB",      "Componentes", 39.90,  22.00,  80, 1],
    [107, "RAM Corsair 8GB",         "Componentes", None,   30.00,  40, 1],   # precio nulo (CRITICO)
    [108, "Motherboard ASUS B450",   "Componentes", 110.00, 78.00,  18, 1],
    [109, "Samsung Galaxy A54",      "Smartphones", 399.00, 300.00, 22, 1],
    [110, "Xiaomi Redmi Note 12",    "Smartphones", 249.00, 180.00, 30, 1],
    [111, "Auriculares HyperX Cloud", None,         79.99,  45.00,  35, 1],   # categoria nula
    [112, "Webcam Logitech C920",    "Perifericos", 89.90,  55.00,  12, 1],
    [103, "Mouse Logitech M170",     "Perifericos", 12.99,  6.50,  120, 1],   # DUPLICADO de id 103
]
prod_highlight = {
    (6, 3),   # precio nulo (id=107)
    (10, 2),  # categoria nula (id=111)
    (12, 0), (12, 1), (12, 2), (12, 3), (12, 4), (12, 5), (12, 6),  # fila duplicada
}
write_sheet(wb, "productos", prod_headers, prod_rows, prod_highlight)

# ---------------------------------------------------------------- ventas (limpia)
ven_headers = ["id_venta", "fecha_venta", "id_cliente", "id_producto",
               "cantidad", "precio_unitario", "descuento", "total_venta", "canal"]
precio_por_prod = {
    101: 850.00, 102: 920.50, 103: 12.99, 104: 45.00, 105: 189.99, 106: 39.90,
    107: 95.00, 108: 110.00, 109: 399.00, 110: 249.00, 111: 79.99, 112: 89.90,
}
prod_ids = list(precio_por_prod.keys())
canales = ["Online", "Tienda"]
start = date(2023, 1, 10)
ven_rows = []
for i in range(1, 51):
    fecha = start + timedelta(days=random.randint(0, 700))  # 2023-2024
    id_cli = random.randint(1, 11)
    id_prod = random.choice(prod_ids)
    cant = random.randint(1, 5)
    punit = precio_por_prod[id_prod]
    desc = round(random.choice([0, 0, 0, 0.05, 0.10]) * punit * cant, 2)
    total = round(punit * cant - desc, 2)
    ven_rows.append([i, fecha, id_cli, id_prod, cant, punit, desc, total,
                     random.choice(canales)])
write_sheet(wb, "ventas", ven_headers, ven_rows, set())

# orden final de hojas
wb.move_sheet("categorias", -(len(wb.sheetnames)))
wb._sheets.sort(key=lambda s: ["clientes", "productos", "ventas", "categorias"].index(s.title))

out = __file__.rsplit("/", 1)[0] + "/Pipeline_ETL_Dataset.xlsx"
wb.save(out)
print("OK ->", out)
print("clientes:", len(cli_rows), "| productos:", len(prod_rows),
      "| ventas:", len(ven_rows), "| categorias:", len(cat_rows))
